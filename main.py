import telebot
import json
import os
import random
import qrcode
import io
import threading
import time
from datetime import datetime, timedelta
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BOT_TOKEN = os.environ.get("BOT_TOKEN")
ADMIN_ID = int(os.environ.get("ADMIN_ID"))

bot = telebot.TeleBot(BOT_TOKEN)
DATA_FILE = "ishlar.json"

# ===== KESH MA'LUMOTLAR =====
_data_cache = None
_last_load_time = 0
CACHE_DURATION = 2

def load_data():
    global _data_cache, _last_load_time
    current_time = time.time()
    if _data_cache is not None and (current_time - _last_load_time) < CACHE_DURATION:
        return _data_cache
    
    if not os.path.exists(DATA_FILE):
        _data_cache = {"ishlar": [], "ishlatilgan_idlar": []}
    else:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            _data_cache = json.load(f)
    _last_load_time = current_time
    return _data_cache

def save_data(data):
    global _data_cache, _last_load_time
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _data_cache = data
    _last_load_time = time.time()

# ===== UNIKAL 4 XONALI ID =====
def yangi_id_yaratish(data):
    ishlatilgan = set(data.get("ishlatilgan_idlar", []))
    while True:
        yangi = random.randint(1000, 9999)
        if yangi not in ishlatilgan:
            ishlatilgan.add(yangi)
            data["ishlatilgan_idlar"] = list(ishlatilgan)
            return yangi

# ===== QR KOD YARATISH =====
def qr_yaratish(ish):
    matn = (
        f"ID: {ish['id']}\n"
        f"Blok: {ish['blok_nomi']}\n"
        f"Mijoz: {ish['mijoz_ismi']}\n"
        f"Tel: {ish['telefon']}\n"
        f"Narx: {ish['narx']} so'm\n"
        f"Sana: {ish['sana']}\n"
        f"Holat: {ish['holat']}"
    )
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=8, border=2)
    qr.add_data(matn)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    return buf

# ===== EXCEL HISOBOT YARATISH =====
def excel_hisobot_yaratish(ishlar, sarlavha):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    headers = ['ID', 'Blok nomi', 'Mijoz ismi', 'Telefon', 'Narx (so\'m)', 'Qabul sanasi', 'Holat', 'Yakunlangan sana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for row_num, ish in enumerate(ishlar, 2):
        row_fill = green_fill if ish['holat'] == 'Yakunlandi' else (yellow_fill if ish['holat'] == 'Jarayonda' else None)
        values = [ish['id'], ish['blok_nomi'], ish['mijoz_ismi'], ish.get('telefon', '-'), ish['narx'], ish['sana'], ish['holat'], ish.get('yakunlangan_sana', '-')]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
    
    col_widths = [6, 25, 20, 15, 15, 18, 14, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ===== KLAVIATURA =====
def main_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(KeyboardButton("➕ Yangi ish qo'shish"))
    kb.add(KeyboardButton("📋 Barcha ishlar"), KeyboardButton("⏳ Jarayondagi ishlar"))
    kb.add(KeyboardButton("🔍 ID bo'yicha qidirish"), KeyboardButton("✅ Ishni yakunlash"))
    kb.add(KeyboardButton("📊 Hisobot"), KeyboardButton("🗑 Ishni o'chirish"))
    return kb

def hisobot_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(KeyboardButton("📅 Bugungi hisobot"))
    kb.add(KeyboardButton("📆 Haftalik hisobot"), KeyboardButton("🗓 10 kunlik hisobot"))
    kb.add(KeyboardButton("📝 Oylik hisobot"), KeyboardButton("✏️ Sana kiritish"))
    kb.add(KeyboardButton("📎 Excel yuklab olish"))
    kb.add(KeyboardButton("🔙 Orqaga"))
    return kb

# Orqaga qaytish uchun alohida klaviatura
def back_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton("🔙 Bekor qilish"))
    return kb

user_states = {}

# ===== XABARLARNI BOSHQARISH =====
def add_message_id(chat_id, message_id):
    if chat_id not in user_states:
        user_states[chat_id] = {"message_ids": [], "hisobot_mode": False, "temp_ids": []}
    if "message_ids" not in user_states[chat_id]:
        user_states[chat_id]["message_ids"] = []
    user_states[chat_id]["message_ids"].append(message_id)

def add_temp_id(chat_id, message_id):
    """Vaqtinchalik xabarlar ID sini saqlash"""
    if chat_id not in user_states:
        user_states[chat_id] = {"message_ids": [], "hisobot_mode": False, "temp_ids": []}
    if "temp_ids" not in user_states[chat_id]:
        user_states[chat_id]["temp_ids"] = []
    user_states[chat_id]["temp_ids"].append(message_id)

def send_and_save(chat_id, text, reply_markup=None, parse_mode=None):
    try:
        msg = bot.send_message(chat_id, text, reply_markup=reply_markup, parse_mode=parse_mode)
        add_message_id(chat_id, msg.message_id)
        return msg
    except Exception as e:
        print(f"Xabar yuborish xatosi: {e}")
        return None

def send_temp_message(chat_id, text, reply_markup=None):
    """Vaqtinchalik xabar yuborish (keyin o'chiriladi)"""
    try:
        msg = bot.send_message(chat_id, text, reply_markup=reply_markup)
        add_temp_id(chat_id, msg.message_id)
        return msg
    except Exception as e:
        print(f"Vaqtinchalik xabar yuborish xatosi: {e}")
        return None

def send_photo_and_save(chat_id, photo, caption=None, reply_markup=None):
    try:
        msg = bot.send_photo(chat_id, photo, caption=caption, reply_markup=reply_markup)
        add_message_id(chat_id, msg.message_id)
        return msg
    except Exception as e:
        print(f"Rasm yuborish xatosi: {e}")
        return None

def send_document_and_save(chat_id, document, visible_file_name, caption=None, reply_markup=None):
    try:
        msg = bot.send_document(chat_id, document, visible_file_name=visible_file_name, caption=caption, reply_markup=reply_markup)
        add_message_id(chat_id, msg.message_id)
        return msg
    except Exception as e:
        print(f"Fayl yuborish xatosi: {e}")
        return None

def delete_temp_messages(chat_id):
    """Vaqtinchalik xabarlarni o'chirish"""
    if chat_id in user_states and "temp_ids" in user_states[chat_id]:
        temp_ids = user_states[chat_id]["temp_ids"].copy()
        user_states[chat_id]["temp_ids"] = []
        for msg_id in temp_ids:
            try:
                bot.delete_message(chat_id, msg_id)
                time.sleep(0.03)
            except:
                pass

def delete_messages_async(chat_id, message_ids):
    if not message_ids:
        return
    def delete():
        for msg_id in message_ids:
            try:
                bot.delete_message(chat_id, msg_id)
                time.sleep(0.03)
            except:
                pass
    threading.Thread(target=delete, daemon=True).start()

def clear_chat_history(chat_id, keep_hisobot=False):
    if chat_id not in user_states:
        return
    
    state = user_states[chat_id]
    
    if keep_hisobot and state.get("hisobot_mode"):
        return
    
    if "message_ids" in state and state["message_ids"]:
        message_ids = state["message_ids"].copy()
        state["message_ids"] = []
        delete_messages_async(chat_id, message_ids)
    
    # Vaqtinchalik xabarlarni ham tozalash
    if "temp_ids" in state and state["temp_ids"]:
        temp_ids = state["temp_ids"].copy()
        state["temp_ids"] = []
        delete_messages_async(chat_id, temp_ids)

def delete_message_later(chat_id, message_id, delay=0):
    if delay > 0:
        def delete():
            time.sleep(delay)
            try:
                bot.delete_message(chat_id, message_id)
            except:
                pass
        threading.Thread(target=delete, daemon=True).start()
    else:
        try:
            bot.delete_message(chat_id, message_id)
        except:
            pass

def delete_multiple_later(chat_id, message_ids, delay=3):
    if not message_ids:
        return
    def delete():
        time.sleep(delay)
        for msg_id in message_ids:
            try:
                bot.delete_message(chat_id, msg_id)
                time.sleep(0.03)
            except:
                pass
    threading.Thread(target=delete, daemon=True).start()

# ===== HISOBOT =====
def hisobot_yaratish(ishlar, sarlavha):
    if not ishlar:
        return f"📊 {sarlavha}\n\n📭 Bu davrda ish yo'q."
    
    jami = len(ishlar)
    yakunlangan = sum(1 for i in ishlar if i["holat"] == "Yakunlandi")
    jarayonda = jami - yakunlangan
    
    jami_summa = 0
    yakunlangan_summa = 0
    for ish in ishlar:
        try:
            narx = int(ish["narx"].replace(" ", "").replace(",", ""))
            jami_summa += narx
            if ish["holat"] == "Yakunlandi":
                yakunlangan_summa += narx
        except:
            pass
    
    return (
        f"📊 {sarlavha}\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"📦 Jami ishlar: {jami} ta\n"
        f"✅ Yakunlangan: {yakunlangan} ta\n"
        f"⏳ Jarayonda: {jarayonda} ta\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"💰 Jami summa: {jami_summa:,} so'm\n"
        f"✅ Yakunlangan summa: {yakunlangan_summa:,} so'm\n"
        f"━━━━━━━━━━━━━━━━"
    )

def ishlar_sanada(ishlar, bosh, oxir):
    natija = []
    for ish in ishlar:
        try:
            sana = datetime.strptime(ish["sana"], "%d.%m.%Y %H:%M")
            if bosh <= sana <= oxir:
                natija.append(ish)
        except:
            pass
    return natija

def show_main_menu(chat_id, text="🏠 Asosiy menyu:"):
    if chat_id in user_states:
        user_states[chat_id]["hisobot_mode"] = False
    delete_temp_messages(chat_id)
    send_and_save(chat_id, text, reply_markup=main_keyboard())

def show_hisobot_menu(chat_id):
    if chat_id in user_states:
        user_states[chat_id]["hisobot_mode"] = True
    send_and_save(chat_id, "📊 Hisobot turini tanlang:", reply_markup=hisobot_keyboard())

# ===== /start =====
@bot.message_handler(commands=["start"])
def start(message):
    if message.from_user.id != ADMIN_ID:
        bot.send_message(message.chat.id, "⛔ Siz admin emassiz.")
        return
    
    clear_chat_history(message.chat.id)
    if message.chat.id in user_states:
        user_states[message.chat.id] = {"message_ids": [], "hisobot_mode": False, "temp_ids": []}
    
    send_and_save(message.chat.id,
        "👋 Xush kelibsiz!\n\n🔧 Mator blok shilish xizmati\nQuyidagi tugmalardan foydalaning:",
        reply_markup=main_keyboard())

def is_admin(message):
    return message.from_user.id == ADMIN_ID

# ===== BEKOR QILISH =====
@bot.message_handler(func=lambda m: is_admin(m) and m.text == "🔙 Bekor qilish")
def bekor_qilish(message):
    clear_chat_history(message.chat.id)
    if message.from_user.id in user_states:
        del user_states[message.from_user.id]
    show_main_menu(message.chat.id)

# ===== ASOSIY MENYU HANDLERLARI =====
@bot.message_handler(func=lambda m: is_admin(m) and m.text == "➕ Yangi ish qo'shish")
def yangi_ish(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "blok_nomi", 
        "data": {}, 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": False,
        "original_menu": "main"
    }
    send_and_save(message.chat.id, "🔧 Blok nomini kiriting:\n(masalan: Toyota Camry 2.4)", reply_markup=back_keyboard())

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📋 Barcha ishlar")
def barcha_ishlar(message):
    clear_chat_history(message.chat.id)
    data = load_data()
    if not data["ishlar"]:
        send_and_save(message.chat.id, "📭 Hozircha ish yo'q.")
        show_main_menu(message.chat.id)
        return
    
    send_and_save(message.chat.id, f"📋 Jami: {len(data['ishlar'])} ta ish")
    for ish in data["ishlar"]:
        holat_emoji = "✅" if ish["holat"] == "Yakunlandi" else "⏳"
        send_and_save(message.chat.id,
            f"🆔 ID: {ish['id']}\n"
            f"🔧 Blok: {ish['blok_nomi']}\n"
            f"👤 Mijoz: {ish['mijoz_ismi']}\n"
            f"📞 Tel: {ish.get('telefon', '-')}\n"
            f"💰 Narx: {ish['narx']} so'm\n"
            f"📅 Sana: {ish['sana']}\n"
            f"{holat_emoji} Holat: {ish['holat']}")
    show_main_menu(message.chat.id)

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "⏳ Jarayondagi ishlar")
def jarayondagi_ishlar(message):
    clear_chat_history(message.chat.id)
    data = load_data()
    jarayon = [i for i in data["ishlar"] if i["holat"] == "Jarayonda"]
    if not jarayon:
        send_and_save(message.chat.id, "✅ Jarayondagi ish yo'q.")
        show_main_menu(message.chat.id)
        return
    
    send_and_save(message.chat.id, f"⏳ Jarayondagi ishlar: {len(jarayon)} ta")
    for ish in jarayon:
        send_and_save(message.chat.id,
            f"🆔 ID: {ish['id']}\n🔧 Blok: {ish['blok_nomi']}\n👤 Mijoz: {ish['mijoz_ismi']}\n"
            f"📞 Tel: {ish.get('telefon', '-')}\n💰 Narx: {ish['narx']} so'm\n📅 Sana: {ish['sana']}\n⏳ Holat: Jarayonda")
    show_main_menu(message.chat.id)

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "🔍 ID bo'yicha qidirish")
def qidirish_boshlash(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "qidirish", 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": False
    }
    # Asosiy menyu bilan birga so'rash
    send_and_save(message.chat.id, "🔍 Ish ID sini kiriting (4 xonali raqam):", reply_markup=main_keyboard())

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "✅ Ishni yakunlash")
def yakunlash_boshlash(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "yakunlash", 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": False
    }
    # Asosiy menyu bilan birga so'rash
    send_and_save(message.chat.id, "✅ Yakunlanmoqchi bo'lgan ish ID sini kiriting:", reply_markup=main_keyboard())

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "🗑 Ishni o'chirish")
def ochirish_boshlash(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "ochirish", 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": False
    }
    # Asosiy menyu bilan birga so'rash
    send_and_save(message.chat.id, "🗑 O'chirmoqchi bo'lgan ish ID sini kiriting:", reply_markup=main_keyboard())

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📊 Hisobot")
def hisobot_menu(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "hisobot_menu", 
        "excel_mode": False, 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": True
    }
    show_hisobot_menu(message.chat.id)

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📎 Excel yuklab olish")
def excel_yuklab_olish(message):
    clear_chat_history(message.chat.id)
    user_states[message.from_user.id] = {
        "bosqich": "hisobot_menu", 
        "excel_mode": True, 
        "message_ids": [], 
        "temp_ids": [],
        "hisobot_mode": True
    }
    show_hisobot_menu(message.chat.id)

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "🔙 Orqaga")
def orqaga(message):
    clear_chat_history(message.chat.id)
    if message.from_user.id in user_states:
        del user_states[message.from_user.id]
    show_main_menu(message.chat.id)

# ===== HISOBOT HANDLERLARI =====
def handle_hisobot(message, days, nom):
    clear_chat_history(message.chat.id, keep_hisobot=True)
    
    data = load_data()
    bosh = datetime.now() - timedelta(days=days)
    ishlar = ishlar_sanada(data["ishlar"], bosh, datetime.now())
    
    state = user_states.get(message.from_user.id, {})
    
    if state.get("excel_mode"):
        if not ishlar:
            send_and_save(message.chat.id, f"📭 {nom} uchun ma'lumot topilmadi.")
        else:
            excel_file = excel_hisobot_yaratish(ishlar, nom)
            send_document_and_save(message.chat.id, excel_file, 
                visible_file_name=f"hisobot_{nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                caption=f"📊 {nom} - {len(ishlar)} ta ish")
        show_main_menu(message.chat.id)
        if message.from_user.id in user_states:
            del user_states[message.from_user.id]
    else:
        send_and_save(message.chat.id, hisobot_yaratish(ishlar, nom))
        show_main_menu(message.chat.id)

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📅 Bugungi hisobot")
def bugungi(message):
    handle_hisobot(message, 1, f"Bugungi hisobot ({datetime.now().strftime('%d.%m.%Y')})")

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📆 Haftalik hisobot")
def haftalik(message):
    handle_hisobot(message, 7, "Haftalik hisobot (7 kun)")

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "🗓 10 kunlik hisobot")
def o_nkunlik(message):
    handle_hisobot(message, 10, "10 kunlik hisobot")

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "📝 Oylik hisobot")
def oylik(message):
    handle_hisobot(message, 30, "Oylik hisobot (30 kun)")

@bot.message_handler(func=lambda m: is_admin(m) and m.text == "✏️ Sana kiritish")
def sana_kiritish(message):
    clear_chat_history(message.chat.id, keep_hisobot=True)
    state = user_states.get(message.from_user.id, {})
    excel_mode = state.get("excel_mode", False)
    
    user_states[message.from_user.id] = {
        "bosqich": "sana_bosh", 
        "data": {},
        "excel_mode": excel_mode,
        "message_ids": [],
        "temp_ids": [],
        "hisobot_mode": True
    }
    send_and_save(message.chat.id, "📅 Boshlanish sanasini kiriting:\nFormat: 01.05.2025")

# ===== BOSQICHMA-BOSQICH =====
@bot.message_handler(func=lambda m: m.from_user.id in user_states)
def handle_steps(message):
    uid = message.from_user.id
    state = user_states.get(uid)
    if not state or "bosqich" not in state:
        return
    
    bosqich = state["bosqich"]
    text = message.text.strip()
    user_msg_id = message.message_id

    # Agar foydalanuvchi "Bekor qilish" tugmasini bossa
    if text == "🔙 Bekor qilish":
        clear_chat_history(message.chat.id)
        del user_states[uid]
        show_main_menu(message.chat.id)
        return

    if bosqich == "qidirish":
        # Avval eski vaqtinchalik xabarlarni o'chirish
        delete_temp_messages(message.chat.id)
        
        try:
            mid = int(text)
            data = load_data()
            ish = next((i for i in data["ishlar"] if i["id"] == mid), None)
            
            if ish:
                # Foydalanuvchi kiritgan raqamni o'chirish
                try:
                    bot.delete_message(message.chat.id, user_msg_id)
                except:
                    pass
                
                holat_emoji = "✅" if ish["holat"] == "Yakunlandi" else "⏳"
                send_and_save(message.chat.id,
                    f"✅ Topildi!\n\n"
                    f"🆔 ID: {ish['id']}\n🔧 Blok: {ish['blok_nomi']}\n👤 Mijoz: {ish['mijoz_ismi']}\n"
                    f"📞 Tel: {ish.get('telefon', '-')}\n💰 Narx: {ish['narx']} so'm\n"
                    f"📅 Sana: {ish['sana']}\n{holat_emoji} Holat: {ish['holat']}")
                qr_buf = qr_yaratish(ish)
                send_photo_and_save(message.chat.id, qr_buf, caption=f"🔲 ID {ish['id']} uchun QR kod")
                show_main_menu(message.chat.id)
                del user_states[uid]
            else:
                # ID topilmadi - vaqtinchalik xabar, menyu o'zgarmaydi
                try:
                    bot.delete_message(message.chat.id, user_msg_id)
                except:
                    pass
                temp_msg = send_temp_message(message.chat.id, f"❌ ID {mid} topilmadi. Boshqa ID kiriting:")
                delete_message_later(message.chat.id, temp_msg.message_id, 3)
                # Holat saqlanadi, menyu o'zgarmaydi
                return
        except:
            try:
                bot.delete_message(message.chat.id, user_msg_id)
            except:
                pass
            temp_msg = send_temp_message(message.chat.id, "⚠️ Faqat raqam kiriting!")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)
            return

    elif bosqich == "yakunlash":
        # Avval eski vaqtinchalik xabarlarni o'chirish
        delete_temp_messages(message.chat.id)
        
        try:
            mid = int(text)
            data = load_data()
            
            for ish in data["ishlar"]:
                if ish["id"] == mid:
                    if ish["holat"] == "Yakunlandi":
                        try:
                            bot.delete_message(message.chat.id, user_msg_id)
                        except:
                            pass
                        temp_msg = send_temp_message(message.chat.id, f"ℹ️ ID {mid} allaqachon yakunlangan.")
                        delete_message_later(message.chat.id, temp_msg.message_id, 3)
                    else:
                        ish["holat"] = "Yakunlandi"
                        ish["yakunlangan_sana"] = datetime.now().strftime("%d.%m.%Y %H:%M")
                        save_data(data)
                        
                        try:
                            bot.delete_message(message.chat.id, user_msg_id)
                        except:
                            pass
                        
                        send_and_save(message.chat.id,
                            f"✅ Ish yakunlandi!\n\n"
                            f"🆔 ID: {ish['id']}\n🔧 Blok: {ish['blok_nomi']}\n👤 Mijoz: {ish['mijoz_ismi']}\n"
                            f"📞 Tel: {ish.get('telefon', '-')}\n💰 Narx: {ish['narx']} so'm\n"
                            f"📅 Qabul: {ish['sana']}\n✅ Yakunlandi: {ish['yakunlangan_sana']}")
                        qr_buf = qr_yaratish(ish)
                        send_photo_and_save(message.chat.id, qr_buf, caption=f"🔲 ID {ish['id']} — Yakunlandi ✅")
                        show_main_menu(message.chat.id)
                    
                    del user_states[uid]
                    return
            
            # ID topilmadi
            try:
                bot.delete_message(message.chat.id, user_msg_id)
            except:
                pass
            temp_msg = send_temp_message(message.chat.id, f"❌ ID {mid} topilmadi. Boshqa ID kiriting:")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)
            return
        except:
            try:
                bot.delete_message(message.chat.id, user_msg_id)
            except:
                pass
            temp_msg = send_temp_message(message.chat.id, "⚠️ Faqat raqam kiriting!")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)
            return

    elif bosqich == "ochirish":
        # Avval eski vaqtinchalik xabarlarni o'chirish
        delete_temp_messages(message.chat.id)
        
        try:
            mid = int(text)
            data = load_data()
            oldin = len(data["ishlar"])
            data["ishlar"] = [i for i in data["ishlar"] if i["id"] != mid]
            
            try:
                bot.delete_message(message.chat.id, user_msg_id)
            except:
                pass
            
            if len(data["ishlar"]) < oldin:
                save_data(data)
                send_and_save(message.chat.id, f"✅ ID {mid} o'chirildi.")
            else:
                temp_msg = send_temp_message(message.chat.id, f"❌ ID {mid} topilmadi.")
                delete_message_later(message.chat.id, temp_msg.message_id, 3)
            
            show_main_menu(message.chat.id)
            del user_states[uid]
        except:
            try:
                bot.delete_message(message.chat.id, user_msg_id)
            except:
                pass
            temp_msg = send_temp_message(message.chat.id, "⚠️ Faqat raqam kiriting!")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)
            show_main_menu(message.chat.id)
            del user_states[uid]

    elif bosqich == "sana_bosh":
        try:
            datetime.strptime(text, "%d.%m.%Y")
            state["data"]["bosh"] = text
            state["bosqich"] = "sana_oxir"
            send_and_save(message.chat.id, "📅 Tugash sanasini kiriting:\nFormat: 31.05.2025")
        except:
            temp_msg = send_temp_message(message.chat.id, "⚠️ Format xato! Masalan: 01.05.2025")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)

    elif bosqich == "sana_oxir":
        try:
            bosh = datetime.strptime(state["data"]["bosh"], "%d.%m.%Y")
            oxir = datetime.strptime(text, "%d.%m.%Y").replace(hour=23, minute=59)
            data = load_data()
            ishlar = ishlar_sanada(data["ishlar"], bosh, oxir)
            
            if state.get("excel_mode"):
                if not ishlar:
                    send_and_save(message.chat.id, f"📭 Hisobot: {state['data']['bosh']} — {text} uchun ma'lumot topilmadi.")
                else:
                    excel_file = excel_hisobot_yaratish(ishlar, f"Hisobot: {state['data']['bosh']} — {text}")
                    send_document_and_save(message.chat.id, excel_file,
                        visible_file_name=f"hisobot_{state['data']['bosh']}_{text}.xlsx",
                        caption=f"📊 Hisobot: {state['data']['bosh']} — {text} - {len(ishlar)} ta ish")
                show_main_menu(message.chat.id)
            else:
                send_and_save(message.chat.id, hisobot_yaratish(ishlar, f"Hisobot: {state['data']['bosh']} — {text}"))
                show_main_menu(message.chat.id)
            
            del user_states[uid]
        except:
            temp_msg = send_temp_message(message.chat.id, "⚠️ Format xato! Masalan: 31.05.2025")
            delete_message_later(message.chat.id, temp_msg.message_id, 3)

    elif bosqich == "blok_nomi":
        state["data"]["blok_nomi"] = text
        state["bosqich"] = "mijoz_ismi"
        send_and_save(message.chat.id, "👤 Mijoz ismini kiriting:", reply_markup=back_keyboard())

    elif bosqich == "mijoz_ismi":
        state["data"]["mijoz_ismi"] = text
        state["bosqich"] = "telefon"
        send_and_save(message.chat.id, "📞 Mijoz telefon raqamini kiriting:", reply_markup=back_keyboard())

    elif bosqich == "telefon":
        state["data"]["telefon"] = text
        state["bosqich"] = "narx"
        send_and_save(message.chat.id, "💰 Kelishilgan narxni kiriting (so'm):", reply_markup=back_keyboard())

    elif bosqich == "narx":
        state["data"]["narx"] = text
        data = load_data()
        yangi_id = yangi_id_yaratish(data)
        sana = datetime.now().strftime("%d.%m.%Y %H:%M")

        yangi_ish_data = {
            "id": yangi_id,
            "blok_nomi": state["data"]["blok_nomi"],
            "mijoz_ismi": state["data"]["mijoz_ismi"],
            "telefon": state["data"]["telefon"],
            "narx": state["data"]["narx"],
            "sana": sana,
            "holat": "Jarayonda",
            "yakunlangan_sana": ""
        }
        data["ishlar"].append(yangi_ish_data)
        save_data(data)
        del user_states[uid]

        send_and_save(message.chat.id,
            f"✅ Ish qo'shildi!\n\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"🆔 ID: {yangi_id}\n🔧 Blok: {yangi_ish_data['blok_nomi']}\n"
            f"👤 Mijoz: {yangi_ish_data['mijoz_ismi']}\n📞 Tel: {yangi_ish_data['telefon']}\n"
            f"💰 Narx: {yangi_ish_data['narx']} so'm\n📅 Sana: {yangi_ish_data['sana']}\n"
            f"⏳ Holat: Jarayonda\n━━━━━━━━━━━━━━━━\n📌 Mijozga bering: ID {yangi_id}")

        qr_buf = qr_yaratish(yangi_ish_data)
        send_photo_and_save(message.chat.id, qr_buf, caption=f"🔲 Mijozga beriladigan QR kod\nID: {yangi_id}")
        show_main_menu(message.chat.id)

# ===== BOTNI ISHGA TUSHIRISH =====
if __name__ == "__main__":
    print("Bot ishga tushdi...")
    bot.remove_webhook()
    bot.infinity_polling(timeout=60, long_polling_timeout=30)
